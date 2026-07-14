import openpyxl, json, re, datetime

wb = openpyxl.load_workbook('/mnt/user-data/uploads/Budget_Alexandre.xlsx', data_only=True)

months_fr = {
    'Janvier':1,'Février':2,'Fevrier':2,'Mars':3,'Avril':4,'Mai':5,'Juin':6,'Juillet':7,
    'Aout':8,'Août':8,'Septembre':9,'Octobre':10,'Novembre':11,'Decembre':12,'Décembre':12
}

def parse_sheet_name(name):
    # e.g. "Budget Alex Janvier 2020" or "Budget Octobre 2020" or "Budget Février 2022"
    m = re.search(r'(' + '|'.join(months_fr.keys()) + r')\s+(\d{4})', name)
    if not m:
        return None
    mon = months_fr[m.group(1)]
    year = int(m.group(2))
    return year, mon

def safe_num(v):
    if isinstance(v, (int, float)):
        return v
    return None

data = {}
skip = {'Liste', 'NewListe'}
for name in wb.sheetnames:
    if name in skip:
        continue
    parsed = parse_sheet_name(name)
    if not parsed:
        continue
    year, mon = parsed
    ws = wb[name]
    summary = {
        'balance_prec': None, 'revenu': None,
        'livretA': None, 'livretA_leandre': None,
        'livretDDS': None, 'livretJoint': None,
        'encours': None, 'previsionnel': None
    }
    items = []
    for row in ws.iter_rows(min_row=1, max_row=10, max_col=6):
        label = row[1].value if len(row) > 1 else None
        if isinstance(label, str):
            l = label.strip().lower()
            if 'balance mois' in l:
                summary['balance_prec'] = safe_num(row[2].value)
            elif l.startswith('revenus'):
                summary['revenu'] = safe_num(row[2].value)
            elif 'livret a' in l and 'prévisionnel' not in l.split('livret a')[1][:3]:
                summary['livretA'] = safe_num(row[2].value)
                # leandre may be in col D/E
                if len(row) > 4 and isinstance(row[3].value, str) and 'andre' in row[3].value.lower():
                    summary['livretA_leandre'] = safe_num(row[4].value)
            elif l.startswith('livret a'):
                summary['livretA'] = safe_num(row[2].value)
                if len(row) > 4 and isinstance(row[3].value, str) and 'andre' in row[3].value.lower():
                    summary['livretA_leandre'] = safe_num(row[4].value)
            elif l.startswith('livret dds') or l.startswith('livret ldds'):
                summary['livretDDS'] = safe_num(row[2].value)
            elif l.startswith('livret joint'):
                summary['livretJoint'] = safe_num(row[2].value)
            elif l == 'en cours':
                summary['encours'] = safe_num(row[2].value)
            elif l == 'prévisionnel' or l == 'previsionnel':
                summary['previsionnel'] = safe_num(row[2].value)
            elif l == 'item':
                continue

    # Now find "Suivi du budget" table rows -- header row with 'Item'
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=6):
        if row[1].value == 'Item':
            header_row = row[0].row
            break
    if header_row:
        for r in range(header_row+1, ws.max_row+1):
            item = ws.cell(row=r, column=2).value
            echeance = ws.cell(row=r, column=3).value
            categorie = ws.cell(row=r, column=4).value
            montant = ws.cell(row=r, column=5).value
            traite = ws.cell(row=r, column=6).value
            if item is None and montant is None and categorie is None:
                continue
            if isinstance(echeance, datetime.datetime):
                echeance = echeance.strftime('%Y-%m-%d')
            items.append({
                'item': item,
                'echeance': echeance,
                'categorie': categorie,
                'montant': safe_num(montant),
                'traite': True if (isinstance(traite,str) and traite.strip().lower()=='oui') else False
            })

    key = f"{year}-{mon:02d}"
    data[key] = {'sheet_name': name, 'summary': summary, 'items': items}

with open('/home/claude/budget/budget_data.json','w',encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False)

print(len(data), 'months extracted')
# print a sample
ks = sorted(data.keys())
print(ks[:3], ks[-3:])
print(json.dumps(data[ks[-1]], ensure_ascii=False, indent=2)[:1500])
